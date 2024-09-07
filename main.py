from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import os

# Create a new workbook and select the active sheet
wb = Workbook()
ws = wb.active

# Set maximum dimensions for cells and images
max_width_px = 300  # Maximum width in pixels
max_height_px = 300  # Maximum height in pixels

# Function to set column width (in Excel's measurement units)
def set_column_width(sheet, col, width_px):
    col_letter = get_column_letter(col)
    sheet.column_dimensions[col_letter].width = width_px / 7  # Roughly 1 Excel unit = 7 pixels

# Function to set row height (in points)
def set_row_height(sheet, row, height_px):
    sheet.row_dimensions[row].height = height_px * 0.75  # 1 point is approximately 0.75 pixels

# Folder containing images
image_folder = "./images"

# Iterate through image files in the folder, starting from B2
start_row = 2
start_col = 2
row = start_row
col = start_col

for filename in os.listdir(image_folder):
    if filename.endswith(".png") or filename.endswith(".jpg") or filename.endswith(".jpeg"):
        img_path = os.path.join(image_folder, filename)

        # Open the image using PIL to get its dimensions
        with PILImage.open(img_path) as pil_img:
            img_width, img_height = pil_img.size
        
        # Calculate scaling factors to fit within the max dimensions
        scale_factor = min(max_width_px / img_width, max_height_px / img_height, 1)
        
        # Adjust the image dimensions based on the scaling factor
        scaled_width = int(img_width * scale_factor)
        scaled_height = int(img_height * scale_factor)

        # Set the column width and row height based on scaled image dimensions
        set_column_width(ws, col, scaled_width)
        set_row_height(ws, row, scaled_height)

        # Open the image using the Image class from openpyxl
        img = Image(img_path)

        # Resize the image to the scaled dimensions
        img.width = scaled_width
        img.height = scaled_height

        # Insert the image into the worksheet at the top-left of the cell
        cell_location = f"{get_column_letter(col)}{row}"
        ws.add_image(img, cell_location)

        # Move to the next row for the next image
        row += 1

# Save the workbook
wb.save("images_in_dynamic_cells.xlsx")
